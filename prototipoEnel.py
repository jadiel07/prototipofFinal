import PySimpleGUI as sg
import openpyxl
from openpyxl import load_workbook
from openpyxl import utils
import win32com.client as win32
import os

# Função de enviar e-mail
def enviar_email(destinatario, assunto, corpo):
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = destinatario
    mail.Subject = assunto
    
    # assinatura
    assinatura_html = r"""
<hr> <!-- Linha horizontal -->
<p><strong>Jadiel Santos de Souza</strong><br>
Estagiário<br>
Gerência de Regulação do Serviço<br>
Diretoria de Regulação da Distribuição e Transmissão<br>
<br>
<img src="C:\Users\BR0528395338\Desktop\python\teste de automatização planilha\prototipos\quase finalizado\assinatura\logo\logoEnel.png" alt="Logotipo Enel Distribuição São Paulo">
<br>
<strong>Enel Distribuição São Paulo</strong><br>
Av. das Nações, 14.401, andar 17 ao 23, Conjunto 1 ao 4, Torre B1 “Aroeira”<br>
Bairro Vila Gertrudes – São Paulo/SP – CEP 04794-000<br>
<a href="mailto:jadiel.souza@enel.com">jadiel.souza@enel.com</a></p>

"""
    mail.HTMLBody = corpo + "<br><br>" + assinatura_html
    
    mail.Send()

# Função para ler as informações da planilha
def ler_planilha(mes):
    meses = {
        'Janeiro': 1,
        'Fevereiro': 2,
        'Março': 3,
        'Abril': 4,
        'Maio': 5,
        'Junho': 6,
        'Julho': 7,
        'Agosto': 8,
        'Setembro': 9,
        'Outubro': 10,
        'Novembro': 11,
        'Dezembro': 12
    }

    mes_numero = meses.get(mes)
    if mes_numero is None:
        print(f'Mês "{mes}" não encontrado.')
        return {}, None  # Retorna os destinatários e o assunto

    coluna_mes = utils.get_column_letter(12 + mes_numero)
    coluna_mes_index = 12 + mes_numero

    print(f'Lendo dados para o mês de {mes}...')
    print(f'Coluna do mês: {coluna_mes}')

    wb = load_workbook('EM ATUALIZAÇÃO - Todas Obrigação Regulares e de Acomp. - Ano 2024.xlsx')
    ws = wb.active

    destinatarios_corpo = {}
    for row in ws.iter_rows(min_row=4, min_col=coluna_mes_index, max_col=coluna_mes_index):
        if row[0].value == 'x':
            destinatario = ws.cell(row=row[0].row, column=11).value
            info_coluna_e = ws.cell(row=row[0].row, column=5).value  # Informação da coluna E e etc
            info_coluna_f = ws.cell(row=row[0].row, column=6).value  
            info_coluna_aa = ws.cell(row=row[0].row, column=27).value  
            
            corpo = f"""
                   <div>
    <p>Prezados (as),</p>
    <p>Dando continuidade ao acompanhamento das informações com envio periódico à ANEEL, pedimos a gentileza,
       de nos encaminhar até dia {info_coluna_f}, as evidências quanto ao cumprimento da obrigação {info_coluna_e}.</p>
    <p>Solicita-se que sejam salvas na pasta de repositório no link>>> <a style="word-wrap: break-word;">{info_coluna_aa}, </a> </p>
    <p>as evidências abaixo listadas: </p>
    <ul>
        <li>Arquivo XML ou ZIP ou XLXS, que foi carregado no sistema da ANEEL; </li>
        <li>Log | PDF | Print com as evidências da confirmação do envio aprovada pela ANEEL.</li>
    </ul>
    <p>É importante mencionar que é de responsabilidade da área de negócio armazenar as evidências de cumprimento da obrigação regular 
       (arquivos submetidos e respectivos comprovantes), sendo que a Regulação solicita cópia dos documentos apenas para comprovação de que o comando regulatório foi cumprido no prazo e com êxito.</p>
       </div>
       """
            
            # Verifica se as informações são válidas e não vazias antes de adicioná-las ao corpo do e-mail
            if destinatario and corpo:
                if destinatario not in destinatarios_corpo:
                    destinatarios_corpo[destinatario] = [] 
                destinatarios_corpo[destinatario].append(corpo)

    print(f'Destinatários e seus corpos encontrados para o mês de {mes}: {destinatarios_corpo}')
    assunto = f'Obrigação regulares de {mes}'

    print(f'Assunto para o mês de {mes}: {assunto}')

    return destinatarios_corpo, assunto

# Tema
sg.theme('Reddit')

# Layout do app
layout = [
    [sg.Text('Envio das Obrigações regulares', size=(30, 1), justification='center')],
    [sg.Button('Janeiro', size=(25, 1))],
    [sg.Button('Fevereiro', size=(25, 1))],
    [sg.Button('Março', size=(25, 1))],
    [sg.Button('Abril', size=(25, 1))],
    [sg.Button('Maio', size=(25, 1))],
    [sg.Button('Junho', size=(25, 1))],
    [sg.Button('Julho', size=(25, 1))],
    [sg.Button('Agosto', size=(25, 1))],
    [sg.Button('Setembro', size=(25, 1))],
    [sg.Button('Outubro', size=(25, 1))],
    [sg.Button('Novembro', size=(25, 1))],
    [sg.Button('Dezembro', size=(25, 1))],
]

# Variável que armazena os e-mails que foram enviados
destinatarios_enviados = set()

window = sg.Window('Envio das Obrigações regulares protótipo', layout, element_justification='c')

while True:
    event, _ = window.read()
    if event == sg.WIN_CLOSED:
        break
    elif event in ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']:
        destinatarios_corpo, assunto = ler_planilha(event)

        if destinatarios_corpo:  # Verificando se há destinatários
            for destinatario, corpos in destinatarios_corpo.items(): 
                for corpo in corpos:  # Envia um e-mail para cada corpo associado ao destinatário
                    enviar_email(destinatario, assunto, corpo)  # Envia o e-mail
                    destinatarios_enviados.add(destinatario)  # Marca o destinatário como enviado 
            sg.popup (f'Todos os e-mails foram enviados com sucesso')

window.close()
