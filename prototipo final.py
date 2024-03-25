import PySimpleGUI as sg
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import openpyxl
from openpyxl import load_workbook
from openpyxl import utils

# Função para enviar e-mail
def enviar_email(destinatario, assunto, corpo):
    remetente = 'aizen15sou@gmail.com'  # Insira o e-mail remetente aqui
    senha = 'kjqj sjtn zbam slmf'  # Insira a senha do e-mail remetente aqui

    mensagem = MIMEMultipart()
    mensagem['From'] = remetente
    mensagem['To'] = destinatario
    mensagem['Subject'] = assunto

    for c in corpo:
        mensagem.attach(MIMEText(c, 'plain'))

    # Conexão com o servidor SMTP do Gmail
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(remetente, senha)
    texto = mensagem.as_string()
    server.sendmail(remetente, destinatario, texto)
    server.quit()


# Função para ler os dados da planilha
def ler_planilha(mes, assunto_global):
    meses = {
        'Janeiro': 1,
        'Fevereiro': 2,
        'Março': 3,
        'Abril': 4,
        'Maio': 5,
        'Junho': 6,
        'Julho': 7,
        'Agosto': 8,
        'Setembro': 9,
        'Outubro': 10,
        'Novembro': 11,
        'Dezembro': 12
    }

    mes_numero = meses.get(mes)
    if mes_numero is None:
        print(f'Mês "{mes}" não encontrado.')
        return {}, None  # Retorna apenas os destinatários e o assunto

    coluna_mes = utils.get_column_letter(12 + mes_numero)
    coluna_mes_index = 12 + mes_numero

    print(f'Lendo dados para o mês de {mes}...')
    print(f'Coluna do mês: {coluna_mes}')

    wb = load_workbook('testpnl.xlsx')
    ws = wb.active

    destinatarios_corpo = {}
    for row in ws.iter_rows(min_row=4, min_col=coluna_mes_index, max_col=coluna_mes_index):
        if row[0].value == 'x':
            destinatario = ws.cell(row=row[0].row, column=11).value
            corpo = ws.cell(row=row[0].row, column=4).value
            if destinatario and corpo:
                if destinatario not in destinatarios_corpo:
                    destinatarios_corpo[destinatario] = []  # Lista para armazenar múltiplos corpos para um destinatário
                destinatarios_corpo[destinatario].append(f'Descrição da atividade: {corpo}')  # Adiciona o corpo à lista

    print(f'Destinatários e seus corpos encontrados para o mês de {mes}: {destinatarios_corpo}')
    assunto = 'Obrigação regular'

    print(f'Assunto para o mês de {mes}: {assunto}')

    return destinatarios_corpo, assunto

# Tema
sg.theme('Reddit')

# Layout do app
layout = [
    [sg.Text('Envio das Obrigações regulares', size=(30, 1), justification='center')],
    [sg.Button('Janeiro', size=(25, 1))],
    [sg.Button('Fevereiro', size=(25, 1))],
    [sg.Button('Março', size=(25, 1))],
    [sg.Button('Abril', size=(25, 1))],
    [sg.Button('Maio', size=(25, 1))],
    [sg.Button('Junho', size=(25, 1))],
    [sg.Button('Julho', size=(25, 1))],
    [sg.Button('Agosto', size=(25, 1))],
    [sg.Button('Setembro', size=(25, 1))],
    [sg.Button('Outubro', size=(25, 1))],
    [sg.Button('Novembro', size=(25, 1))],
    [sg.Button('Dezembro', size=(25, 1))],
]

# Variável que armazena os e-mails que foram enviados
destinatarios_enviados = set()

# Janela
window = sg.Window('Envio das Obrigações regulares protótipo', layout, element_justification='c')

while True:
    event, _ = window.read()
    if event == sg.WIN_CLOSED:
        break
    elif event in ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']:
        destinatarios_corpo, assunto = ler_planilha(event, 'assunto global')

        if destinatarios_corpo:  # Verificando se há destinatários
            for destinatario, corpos in destinatarios_corpo.items():  # Loop sobre destinatários e seus corpos
                for corpo in corpos:  # Envia um e-mail para cada corpo associado ao destinatário
                    enviar_email(destinatario, assunto, corpo)  # Envia o e-mail
                    destinatarios_enviados.add(destinatario)  # Marca o destinatário como enviado
            sg.popup(f'Todos os e-mails foram enviados com sucesso!')
        else:
            sg.popup(f'Erro: Nenhum destinatário encontrado para o mês {event}!')


window.close()
